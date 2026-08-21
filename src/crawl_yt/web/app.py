"""Server-rendered local dashboard over existing crawl-yt services."""

from __future__ import annotations

from pathlib import Path
from typing import Any
from urllib.parse import quote, urlencode
import json
import re
from io import BytesIO
from datetime import datetime, timezone

from fastapi import FastAPI, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from ..collectors.channel_collector import ChannelCrawlService
from ..collectors.video_metadata import VideoMetadataService
from ..collectors.ytdlp_channel_video import YtDlpChannelVideoProvider
from ..collectors.ytdlp_video_metadata import YtDlpVideoMetadataProvider
from ..database.models import OperationalBudget
from ..database.repository import TranscriptRepository, VideoRepository
from ..discovery.channel_discovery import DiscoveryService
from ..discovery.ytdlp_provider import YtDlpDiscoveryProvider
from ..operations.planner import OperationalPlanner, WorkPlanExecutor
from ..operations.video_scoring import VideoScoringService
from ..transcripts.provider import TranscriptService
from ..transcripts.ytdlp_provider import YtDlpTranscriptProvider
from ..scoring_lifecycle import ChannelScoringLifecycle
from ..crawl_policy import CrawlPriorityPolicy
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).parent
templates = Jinja2Templates(directory=str(ROOT / "templates"))
templates.env.globals["quote"] = quote


def create_app(
    repository: VideoRepository | None = None,
    discovery_provider: Any | None = None,
    video_provider: Any | None = None,
    metadata_provider: Any | None = None,
    transcript_provider: Any | None = None,
) -> FastAPI:
    database = repository or TranscriptRepository()
    app = FastAPI(title="crawl-yt dashboard")
    app.mount("/static", StaticFiles(directory=str(ROOT / "static")), name="static")

    @app.exception_handler(HTTPException)
    async def friendly_http_error(request: Request, exc: HTTPException) -> HTMLResponse:
        return HTMLResponse(str(exc.detail), status_code=exc.status_code)

    @app.exception_handler(Exception)
    async def friendly_error(request: Request, exc: Exception) -> HTMLResponse:
        if isinstance(exc, HTTPException):
            return HTMLResponse(str(exc.detail), status_code=exc.status_code)
        return templates.TemplateResponse(
            request=request,
            name="result.html",
            context={"title": "Operation failed", "result": {"error": str(exc)}},
            status_code=502,
        )

    def render(request: Request, template: str, **values: Any) -> HTMLResponse:
        return templates.TemplateResponse(request=request, name=template, context=values)

    @app.get("/health")
    def health() -> dict[str, str]:
        return {"status": "ok"}

    @app.get("/", response_class=HTMLResponse)
    def dashboard(request: Request) -> HTMLResponse:
        stats = {
            "channels": database.count_channels(),
            "videos": database.count_videos(),
            "metadata_enriched": database.count_enriched_videos(),
            "metadata_pending": database.count_videos_needing_enrichment(),
            "transcripts": database.count_transcripts(),
            "with_transcript": database.count_videos_with_transcripts(),
            "without_transcript": database.count_videos_without_transcripts(),
            "discovery_relationships": database.count_discovery_relationships(),
            "channel_tiers": database.count_channels_by_score_tier(),
            "video_tiers": database.count_video_score_tiers(),
            "work_plans": database.work_plan_status_counts(),
            "crawl_due": database.count_channels_due_for_crawl(),
            "failing": database.count_failing_channels(),
        }
        return render(request, "dashboard.html", title="Dashboard", stats=stats)

    @app.get("/discovery", response_class=HTMLResponse)
    def discovery_page(request: Request) -> HTMLResponse:
        return render(request, "form.html", title="Discovery", form="discovery", history=database.list_discovery_keyword_summaries())

    @app.post("/discovery", response_class=HTMLResponse)
    def discovery_action(request: Request, keyword: str = Form(...), limit: int = Form(...), dry_run: bool = Form(False)) -> HTMLResponse:
        if not keyword.strip() or not 1 <= limit <= 1000:
            raise HTTPException(400, "Keyword and limit 1-1000 are required")
        provider = discovery_provider or YtDlpDiscoveryProvider()
        report = DiscoveryService(provider, database).discover(keyword.strip(), limit, dry_run=dry_run)
        return render(
            request,
            "discovery_result.html",
            title="Discovery result",
            report=report,
            channels=report.channels,
            new_channel_ids=set(report.new_channel_ids),
            dry_run=dry_run,
            keyword=keyword.strip(),
            channels_scored=report.channels_scored,
            scoring_failures=report.scoring_failures,
        )

    @app.get("/discovery/keywords/{keyword}", response_class=HTMLResponse)
    def discovery_keyword_detail(request: Request, keyword: str, page: int = 1, per_page: int = 50) -> HTMLResponse:
        page = max(page, 1)
        per_page = per_page if per_page in {25, 50, 100} else 50
        rows = database.list_channels_for_discovery_keyword(keyword, per_page, (page - 1) * per_page)
        total = database.count_channels_for_discovery_keyword(keyword)
        return render(request, "discovery_keyword.html", title=f"Discovery: {keyword}", keyword=keyword, rows=rows, page=page, per_page=per_page, total=total)

    @app.get("/channels", response_class=HTMLResponse)
    def channels(request: Request, page: int = 1, per_page: str = "50", search: str | None = None, tier: str | None = None, keyword: str | None = None, min_videos_per_week: str | None = None, min_score: str | None = None, sort: str = "score", message: str | None = None) -> HTMLResponse:
        try:
            requested_page_size = int(per_page)
        except (TypeError, ValueError):
            requested_page_size = 50
        per_page = requested_page_size if requested_page_size in {25, 50, 100} else 50
        page = max(page, 1)
        min_videos_per_week = _parse_optional_float(min_videos_per_week, "minimum videos/week", minimum=0)
        min_score = _parse_optional_float(min_score, "minimum score", minimum=0, maximum=100)
        if sort not in {"score", "cadence", "subscribers"}:
            sort = "score"
        total = database.count_channels_page(search, tier, keyword, min_videos_per_week, min_score)
        page_count = max(1, (total + per_page - 1) // per_page)
        page = min(page, page_count)
        rows = database.list_channels_page(per_page, (page - 1) * per_page, search, tier, keyword, min_videos_per_week, sort, min_score)
        query = {key: value for key, value in {"per_page": per_page, "search": search, "tier": tier, "keyword": keyword, "min_videos_per_week": min_videos_per_week, "min_score": min_score, "sort": sort}.items() if value not in (None, "")}
        previous_url = "/channels?" + urlencode({**query, "page": page - 1}) if page > 1 else None
        next_url = "/channels?" + urlencode({**query, "page": page + 1}) if page < page_count else None
        return render(request, "list.html", title="Channels", page_kind="channels", rows=rows, page=page, page_count=page_count, per_page=per_page, total=total, search=search, tier=tier, keyword=keyword, keyword_options=database.list_discovery_keywords(), min_videos_per_week=min_videos_per_week, min_score=min_score, sort=sort, message=message, previous_url=previous_url, next_url=next_url, crawl_interval_labels={"high": "3 days", "medium": "7 days", "low": "14 days", "unscored": "1 day"})

    @app.get("/channels/export.xlsx")
    def export_channels(search: str | None = None, tier: str | None = None,
                        keyword: str | None = None, min_videos_per_week: str | None = None,
                        min_score: str | None = None, sort: str = "score") -> StreamingResponse:
        min_videos_per_week = _parse_optional_float(min_videos_per_week, "minimum videos/week", minimum=0)
        min_score = _parse_optional_float(min_score, "minimum score", minimum=0, maximum=100)
        if min_score is None:
            min_score = 60
        if sort not in {"score", "cadence", "subscribers"}:
            sort = "score"
        try:
            rows = database.list_channels_for_export(search=search, tier=tier, keyword=keyword, min_videos_per_week=min_videos_per_week, min_score=min_score, sort=sort)
        except ValueError as error:
            raise HTTPException(400, str(error)) from error
        if not rows:
            raise HTTPException(400, "No channels match the export filters.")
        workbook = _channels_workbook(rows)
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        slug = re.sub(r"[^a-z0-9]+", "-", (keyword or "channels").strip().lower()).strip("-") or "channels"
        filename = f"crawl-yt-{slug}-score-{min_score:g}.xlsx"
        return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    @app.get("/channels/{channel_id}", response_class=HTMLResponse)
    def channel_detail(request: Request, channel_id: str) -> HTMLResponse:
        channel = database.get_channel(channel_id)
        if channel is None:
            raise HTTPException(404, "Channel not found")
        discoveries = database.list_discoveries_for_channel(channel_id)
        score = database.get_channel_score(channel_id)
        state = database.get_channel_crawl_state(channel_id)
        return render(request, "detail.html", title=channel.title, kind="channel", channel=channel,
                      score=score, score_view=_channel_score_view(score), state=state, state_view=_crawl_state_view(state, score),
                      discoveries=[{"keyword": item.keyword, "source": item.source, "discovered_at": _format_datetime(item.discovered_at)} for item in discoveries], videos=database.list_videos_for_channel(channel_id, 20))

    @app.post("/channels/{channel_id}/score")
    def channel_score(request: Request, channel_id: str) -> RedirectResponse:
        from ..discovery.channel_scoring import ChannelScoringService
        if database.get_channel(channel_id) is None:
            raise HTTPException(404, "Channel not found")
        ChannelScoringService(database).score_channel(channel_id)
        return RedirectResponse(f"/channels/{channel_id}", status_code=303)

    @app.post("/channels/score-unscored")
    def score_unscored_channels(limit: int = Form(...)) -> RedirectResponse:
        if limit < 1:
            raise HTTPException(400, "limit must be positive")
        result = ChannelScoringLifecycle(database).score_unscored_channels(limit)
        message = f"Channels scored: {result.channels_scored}; scoring failures: {len(result.scoring_failures)}"
        return RedirectResponse(f"/channels?message={quote(message)}", status_code=303)

    @app.post("/channels/{channel_id}/crawl")
    def channel_crawl(channel_id: str, full: bool = Form(False)) -> RedirectResponse:
        if database.get_channel(channel_id) is None:
            raise HTTPException(404, "Channel not found")
        service = ChannelCrawlService(video_provider or YtDlpChannelVideoProvider(), database)
        service.crawl(channel_id, full=full)
        return RedirectResponse(f"/channels/{channel_id}", status_code=303)

    @app.post("/channels/{channel_id}/crawl-full")
    def channel_crawl_full(channel_id: str, confirm: bool = Form(False)) -> RedirectResponse:
        if database.get_channel(channel_id) is None:
            raise HTTPException(404, "Channel not found")
        if not confirm:
            raise HTTPException(400, "Full crawl requires explicit confirmation")
        service = ChannelCrawlService(video_provider or YtDlpChannelVideoProvider(), database)
        service.crawl(channel_id, full=True)
        return RedirectResponse(f"/channels/{channel_id}", status_code=303)

    @app.get("/videos", response_class=HTMLResponse)
    def videos(request: Request, page: int = 1, per_page: int = 50, channel_id: str | None = None, tier: str | None = None, metadata_pending: bool = False, transcript_pending: bool = False) -> HTMLResponse:
        per_page = min(max(per_page, 1), 100)
        page = max(page, 1)
        total = database.count_videos_page(channel_id, tier, metadata_pending, transcript_pending)
        rows = database.list_videos_page(per_page, (page - 1) * per_page, channel_id, tier, metadata_pending, transcript_pending)
        return render(request, "list.html", title="Videos", page_kind="videos", rows=rows, page=page, per_page=per_page, total=total)

    @app.get("/videos/{video_id}", response_class=HTMLResponse)
    def video_detail(request: Request, video_id: str) -> HTMLResponse:
        video = database.get_video(video_id)
        if video is None:
            raise HTTPException(404, "Video not found")
        transcripts = database.list_transcripts_for_video(video_id)
        video_score = database.get_video_score(video_id)
        return render(request, "detail.html", title=video.title, kind="video", video=video,
                      score=video_score, video_score_view=_video_score_view(video_score), transcripts=transcripts)

    @app.get("/videos/{video_id}/transcripts/{language}/{source}", response_class=HTMLResponse)
    def transcript_detail(request: Request, video_id: str, language: str, source: str) -> HTMLResponse:
        if database.get_video(video_id) is None:
            raise HTTPException(404, "Video not found")
        transcript = database.get_transcript(video_id, language, source)
        if transcript is None:
            raise HTTPException(404, "Transcript not found")
        return render(request, "transcript.html", title=f"Transcript · {video_id}", video_id=video_id, transcript=transcript, segments=_format_segments(transcript.segments))

    @app.post("/videos/{video_id}/score")
    def video_score(video_id: str) -> RedirectResponse:
        if database.get_video(video_id) is None:
            raise HTTPException(404, "Video not found")
        VideoScoringService(database).score_video(video_id)
        return RedirectResponse(f"/videos/{video_id}", status_code=303)

    @app.post("/videos/{video_id}/enrich")
    def video_enrich(video_id: str) -> RedirectResponse:
        if database.get_video(video_id) is None:
            raise HTTPException(404, "Video not found")
        VideoMetadataService(metadata_provider or YtDlpVideoMetadataProvider(), database).enrich(video_id)
        return RedirectResponse(f"/videos/{video_id}", status_code=303)

    @app.post("/videos/{video_id}/transcript")
    def video_transcript(video_id: str, language: str | None = Form(None)) -> RedirectResponse:
        if database.get_video(video_id) is None:
            raise HTTPException(404, "Video not found")
        TranscriptService(transcript_provider or YtDlpTranscriptProvider(), database).transcript(video_id, language)
        return RedirectResponse(f"/videos/{video_id}", status_code=303)

    @app.get("/work-plans", response_class=HTMLResponse)
    def work_plans(request: Request, page: int = 1, per_page: int = 20) -> HTMLResponse:
        per_page = min(max(per_page, 1), 100)
        page = max(page, 1)
        rows = database.list_work_plans_page(per_page, (page - 1) * per_page)
        return render(request, "list.html", title="Work Plans", page_kind="work-plans", rows=rows, page=page, per_page=per_page, total=database.count_work_plans())

    @app.get("/work-plans/new", response_class=HTMLResponse)
    def new_work_plan(request: Request) -> HTMLResponse:
        return render(request, "new_work_plan.html", title="Create Work Plan")

    @app.post("/work-plans", response_class=HTMLResponse)
    def create_work_plan(request: Request, max_crawls: int = Form(...), max_enrichments: int = Form(...), max_transcripts: int = Form(...)) -> RedirectResponse:
        if min(max_crawls, max_enrichments, max_transcripts) < 0:
            raise HTTPException(400, "Budgets must be non-negative")
        plan = OperationalPlanner(database).plan(OperationalBudget(max_crawls, max_enrichments, max_transcripts))
        return RedirectResponse(f"/work-plans/{plan.id}", status_code=303)

    @app.get("/work-plans/{plan_id}", response_class=HTMLResponse)
    def work_plan_detail(request: Request, plan_id: int) -> HTMLResponse:
        plan = database.get_work_plan(plan_id)
        if plan is None:
            raise HTTPException(404, "Work plan not found")
        return render(request, "detail.html", title=f"Work Plan {plan_id}", kind="work-plan", plan=plan, items=database.list_work_items(plan_id))

    @app.post("/work-plans/{plan_id}/execute")
    def execute_work_plan(request: Request, plan_id: int, max_items: int | None = Form(None), retry_failed: bool = Form(False)) -> RedirectResponse:
        if database.get_work_plan(plan_id) is None:
            raise HTTPException(404, "Work plan not found")
        if max_items is None or max_items < 1:
            raise HTTPException(400, "max_items must be explicit and positive")
        from ..collectors.channel_collector import ChannelCrawlService
        from ..collectors.video_metadata import VideoMetadataService
        from ..transcripts.provider import TranscriptService
        executor = WorkPlanExecutor(database, ChannelCrawlService(video_provider or YtDlpChannelVideoProvider(), database), VideoMetadataService(metadata_provider or YtDlpVideoMetadataProvider(), database), TranscriptService(transcript_provider or YtDlpTranscriptProvider(), database))
        executor.execute(plan_id, max_items, retry_failed)
        return RedirectResponse(f"/work-plans/{plan_id}", status_code=303)

    return app


def _format_segments(segments: list[dict[str, Any]]) -> list[dict[str, str]]:
    formatted: list[dict[str, str]] = []
    for segment in segments:
        try:
            seconds = float(segment.get("start", segment.get("start_time", 0)))
        except (TypeError, ValueError):
            seconds = 0.0
        total_millis = max(0, int(round(seconds * 1000)))
        hours, remainder = divmod(total_millis, 3_600_000)
        minutes, remainder = divmod(remainder, 60_000)
        secs, millis = divmod(remainder, 1000)
        formatted.append({"timestamp": f"{hours:02d}:{minutes:02d}:{secs:02d}.{millis:03d}", "text": str(segment.get("text", ""))})
    return formatted


def _format_datetime(value: datetime | None) -> str | None:
    if value is None:
        return None
    current = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    return current.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _channel_score_view(score: Any | None) -> dict[str, Any] | None:
    if score is None:
        return None
    reasons = score.reasons if isinstance(score.reasons, dict) else {}
    rate_30 = reasons.get("videos_per_week_30d", score.videos_per_week_30d)
    rate_90 = reasons.get("videos_per_week_90d", score.videos_per_week_90d)
    coverage_30 = reasons.get("observation_coverage_30d", False)
    return {
        "overall": score.score,
        "tier": str(score.tier).upper(),
        "components": [("Topic Relevance", score.relevance_score), ("Upload cadence (Activity)", score.activity_score), ("Performance (Traction)", score.traction_score), ("Confidence", score.confidence_score)],
        "version": score.scoring_version,
        "scored_at": _format_datetime(score.scored_at),
        "notes": list(reasons.get("notes", [])),
        "videos_per_week_30d": rate_30 if coverage_30 else None,
        "videos_per_week_90d": rate_90 if coverage_30 else None,
        "observation_coverage_30d": coverage_30,
        "cadence_fit": reasons.get("cadence_fit", "unknown"),
        "consistency": reasons.get("consistency", "unknown"),
        "maturity": reasons.get("score_maturity", "preliminary"),
        "median_views": reasons.get("median_enriched_views"),
        "subscriber_count": reasons.get("subscriber_count"),
        "view_subscriber_ratio": reasons.get("view_subscriber_ratio"),
    }


def _video_score_view(score: Any | None) -> dict[str, Any] | None:
    if score is None:
        return None
    try:
        reasons = json.loads(score.reason_json)
    except (TypeError, ValueError):
        reasons = {}
    return {
        "overall": score.score,
        "tier": str(score.tier).upper(),
        "components": [("Recency", score.recency_score), ("Channel", score.channel_score), ("Traction", score.traction_score), ("Metadata priority", score.metadata_priority), ("Transcript priority", score.transcript_priority), ("Confidence", score.confidence_score)],
        "version": score.scoring_version,
        "scored_at": _format_datetime(score.scored_at),
        "notes": list(reasons.get("notes", [])) if isinstance(reasons, dict) else [],
    }


def _crawl_state_view(state: Any | None, score: Any | None = None) -> dict[str, Any]:
    policy = CrawlPriorityPolicy()
    tier = getattr(score, "tier", None) if score is not None else None
    if state is None:
        return {
            "tier": tier or "unscored",
            "recommended_interval": _format_interval(policy.interval_for(tier)),
            "last_crawl": None,
            "last_success": None,
            "last_error": None,
            "next_crawl": None,
            "total_crawls": 0,
            "failures": 0,
        }
    return {
        "tier": tier or "unscored",
        "recommended_interval": _format_interval(policy.interval_for(tier)),
        "last_crawl": _format_datetime(state.last_crawl_completed_at),
        "last_success": _format_datetime(state.last_success_at),
        "last_error": state.last_error,
        "next_crawl": _format_datetime(state.next_crawl_at),
        "total_crawls": state.total_crawls,
        "failures": state.consecutive_failures,
    }


def _format_interval(value) -> str:
    days = value.days
    return f"{days} day" if days == 1 else f"{days} days"


def _parse_optional_float(
    value: str | None,
    field_name: str,
    minimum: float | None = None,
    maximum: float | None = None,
) -> float | None:
    if value is None or not value.strip():
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError) as error:
        raise HTTPException(400, f"{field_name} must be a number") from error
    if minimum is not None and parsed < minimum:
        if maximum is not None:
            raise HTTPException(400, f"{field_name} must be between {minimum:g} and {maximum:g}")
        raise HTTPException(400, f"{field_name} cannot be negative")
    if maximum is not None and parsed > maximum:
        raise HTTPException(400, f"{field_name} must be between {minimum:g} and {maximum:g}")
    return parsed


def _channels_workbook(rows: list[dict[str, object]]) -> Workbook:
    headers = ["Channel", "Channel URL", "Channel ID", "Overall Score", "Tier", "Videos/Week 30d", "Videos/Week 90d", "Cadence Fit", "Subscribers", "Median Video Views", "Views / Subscribers", "Discovery Keywords", "Score Maturity", "Scoring Version", "Last Crawl", "Next Crawl"]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Channels"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        try:
            reasons = json.loads(row.get("reason_json") or "{}")
        except (TypeError, ValueError):
            reasons = {}
        url = row.get("channel_url") or f"https://www.youtube.com/channel/{row.get('channel_id')}"
        values = [row.get("title"), url, row.get("channel_id"), row.get("score"), row.get("tier"), row.get("videos_per_week_30d"), row.get("videos_per_week_90d"), row.get("cadence_fit"), row.get("subscriber_count"), reasons.get("median_enriched_views"), reasons.get("view_subscriber_ratio"), row.get("discovery_keywords"), reasons.get("score_maturity"), row.get("scoring_version"), row.get("last_success_at"), row.get("next_crawl_at")]
        sheet.append(values)
        link = sheet.cell(sheet.max_row, 2)
        link.hyperlink = str(url)
        link.style = "Hyperlink"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:P{sheet.max_row}"
    widths = [28, 48, 18, 15, 12, 17, 17, 16, 15, 20, 20, 28, 18, 18, 22, 22]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    return workbook
