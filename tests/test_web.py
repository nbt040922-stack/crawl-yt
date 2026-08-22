"""Network-free Phase 3A web adapter tests."""

from __future__ import annotations

import tempfile
import unittest
from io import BytesIO
from datetime import datetime, timedelta, timezone
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.crawl_yt.collectors.video_metadata import VideoMetadata
from src.crawl_yt.collectors.channel_metadata import ChannelMetadata
from src.crawl_yt.database.models import Channel, ChannelScore, Transcript, Video, VideoScore
from src.crawl_yt.database.repository import TranscriptRepository
from src.crawl_yt.discovery.channel_discovery import ChannelVerification, DiscoveryBatch
from src.crawl_yt.transcripts.provider import TranscriptData
from src.crawl_yt.web.app import create_app


def _dated_videos(channel: Channel, count: int = 20) -> list[Video]:
    now = datetime.now(timezone.utc)
    return [Video(f"{channel.channel_id}-{index}", channel.channel_id, "sample", now, published_at=now - timedelta(days=index)) for index in range(count)]


class DiscoveryFake:
    def __init__(self):
        self.calls = []

    def search(self, keyword, limit):
        self.calls.append((keyword, limit))
        self.keyword = keyword
        return DiscoveryBatch(2, [Channel("UC1", "One", channel_url="https://youtube.com/channel/UC1"), Channel("UC3", "Two", channel_url="https://youtube.com/channel/UC3")], "fake")

    def verify(self, channel, sample_size=20):
        return ChannelVerification(channel, [f"{self.keyword} planning"] * sample_size, _dated_videos(channel, sample_size))


class EmptyDiscoveryFake:
    def search(self, keyword, limit):
        return DiscoveryBatch(0, [], "fake")


class RelevanceDiscoveryFake:
    def search(self, keyword, limit):
        self.keyword = keyword
        return DiscoveryBatch(2, [Channel("UCA", "Accepted"), Channel("UCR", "Rejected")], "fake")

    def verify(self, channel, sample_size=20):
        count = 20 if channel.channel_id == "UCA" else 1
        return ChannelVerification(channel, [f"{self.keyword} planning"] * count + ["Unrelated topic"] * (sample_size - count), _dated_videos(channel, sample_size))


class ModeDiscoveryFake:
    def __init__(self, matches):
        self.matches = matches
        self.calls = []

    def search(self, keyword, limit):
        self.calls.append((keyword, limit))
        self.keyword = keyword
        return DiscoveryBatch(100, [Channel("UCMODE", "Mode Candidate")], "fake")

    def verify(self, channel, sample_size=20):
        titles = [f"{self.keyword} planning" if index % 2 == 0 else "Living alone tips" for index in range(self.matches)]
        return ChannelVerification(channel, titles + ["Unrelated topic"] * (sample_size - self.matches), _dated_videos(channel, sample_size))


class EmptySampleDiscoveryFake:
    def search(self, keyword, limit):
        return DiscoveryBatch(1, [Channel("UCEMPTY", "Empty Sample")], "fake")

    def verify(self, channel, sample_size=20):
        return ChannelVerification(channel, [])


class ProfileDiscoveryFake:
    def __init__(self):
        self.search_calls = []
        self.verify_calls = []

    def search(self, keyword, limit):
        self.search_calls.append((keyword, limit))
        return DiscoveryBatch(1, [Channel("UCPROFILE", "Senior Life Solo Path")], "fake")

    def verify(self, channel, sample_size=20):
        self.verify_calls.append((channel.channel_id, sample_size))
        titles = ["Why I enjoy living alone after 65"] * 13 + ["Kitchen tools"] * 7
        return ChannelVerification(channel, titles, _dated_videos(channel, sample_size))


class MultiQueryDiscoveryFake:
    def search(self, keyword, limit):
        if keyword == "broken query":
            raise RuntimeError("search unavailable")
        channels = {
            "retirement": [
                Channel("UC-MATCH", "Matched channel"),
                Channel("UC-REJECT", "Rejected channel"),
            ],
            "retirement income": [
                Channel("UC-MATCH", "Matched channel"),
                Channel("UC-SECOND", "Second matched channel"),
            ],
        }
        batch = channels[keyword][:limit]
        return DiscoveryBatch(len(batch), batch, "fake")

    def verify(self, channel, sample_size=20):
        titles = (
            ["Unrelated topic"] * sample_size
            if channel.channel_id == "UC-REJECT"
            else ["Retirement planning advice"] * sample_size
        )
        return ChannelVerification(channel, titles, _dated_videos(channel, sample_size))


class VideoFake:
    def __init__(self, video=None):
        self.calls = []
        self.video = video

    def iterate_videos(self, channel_id, limit=None):
        self.calls.append((channel_id, limit))
        return [self.video] if self.video else []


class BatchFailureFake(VideoFake):
    def iterate_videos(self, channel_id, limit=None):
        self.calls.append((channel_id, limit))
        if channel_id == "UC1":
            raise RuntimeError("provider failed")
        return []


class MetadataFake:
    def __init__(self):
        self.calls = []

    def fetch(self, video_id, webpage_url=None):
        self.calls.append((video_id, webpage_url))
        return VideoMetadata(video_id, "fake", title="Enriched title", view_count=42)


class TranscriptFake:
    def __init__(self):
        self.calls = []

    def fetch(self, video_id, webpage_url, preferred_languages):
        self.calls.append((video_id, webpage_url, preferred_languages))
        return TranscriptData(video_id, "en", "fake", "Fetched text", [{"start": 1.2, "text": "Fetched text"}])


class WebTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.repository = TranscriptRepository(Path(self.temp.name) / "web.db")
        self.repository.upsert_channel(Channel("UC1", "One", subscriber_count=10))
        self.repository.upsert_channel(Channel("UC2", "Two", subscriber_count=20))
        now = datetime.now(timezone.utc)
        self.repository.record_discovery("UC1", "retirement", "seed", now.replace(hour=8))
        self.repository.record_discovery("UC1", " social security ", "seed", now.replace(hour=9))
        self.repository.record_discovery("UC2", "RETIREMENT", "seed", now.replace(hour=10))
        self.repository.upsert_video(Video("v1", "UC1", "Video one", now))
        self.repository.upsert_video(Video("v2", "UC1", "Video two", now))
        self.repository.upsert_transcript(Transcript("v1", "en", "youtube_auto", "Police arrived.", [{"start": 12.4, "text": "Police arrived."}], now, now))
        self.client = TestClient(create_app(repository=self.repository))

    def tearDown(self) -> None:
        self.temp.cleanup()

    def test_health_and_dashboard(self) -> None:
        self.assertEqual(self.client.get("/health").json(), {"status": "ok"})
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Channels", response.text)
        self.assertIn("Videos", response.text)

    def test_pages_and_pagination(self) -> None:
        self.assertEqual(self.client.get("/channels?page=1&per_page=1").status_code, 200)
        self.assertIn("One", self.client.get("/channels").text)
        self.assertEqual(self.client.get("/videos?page=1&per_page=1").status_code, 200)
        self.assertIn("Video one", self.client.get("/videos").text)
        self.assertEqual(self.client.get("/work-plans").status_code, 200)

    def test_full_crawl_batch_creation_uses_filters_and_confirmation(self) -> None:
        response = self.client.post("/channels/full-crawl-batch", data={"limit": "1", "keyword": "retirement"})
        self.assertEqual(response.status_code, 400)
        provider = VideoFake()
        client = TestClient(create_app(repository=self.repository, video_provider=provider))
        response = client.post("/channels/full-crawl-batch", data={"limit": "1", "keyword": "retirement", "confirmation": "true"}, follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        self.assertEqual(provider.calls, [])
        batch = self.repository.get_crawl_batch(1)
        self.assertEqual(batch.candidate_count, 2)
        self.assertEqual(batch.selected_count, 1)
        self.assertEqual(self.repository.list_crawl_batch_items(1)[0].channel_id, "UC1")

    def test_full_crawl_batch_creation_limit_and_pagination_are_independent(self) -> None:
        repo = TranscriptRepository(Path(self.temp.name) / "batch-many.db")
        for number in range(81):
            channel_id = f"B{number:03d}"
            repo.upsert_channel(Channel(channel_id, f"Batch {number:03d}"))
            repo.record_discovery(channel_id, "retirement", "seed")
        client = TestClient(create_app(repository=repo))
        response = client.post("/channels/full-crawl-batch?page=2&per_page=25", data={"limit": "50", "keyword": "retirement", "confirmation": "true"}, follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        batch = repo.get_crawl_batch(1)
        self.assertEqual((batch.candidate_count, batch.selected_count), (81, 50))
        self.assertEqual(repo.list_crawl_batch_items(1)[0].channel_id, "B000")

    def test_full_crawl_batch_run_next_is_chunked_and_failures_continue(self) -> None:
        provider = VideoFake(Video("batch-video", "UC1", "Batch video", datetime.now(timezone.utc)))
        client = TestClient(create_app(repository=self.repository, video_provider=provider))
        created = client.post("/channels/full-crawl-batch", data={"limit": "2", "confirmation": "true"}, follow_redirects=False)
        self.assertEqual(created.status_code, 303)
        detail = client.post("/crawl-batches/1/run-next", data={"chunk_size": "1"}, follow_redirects=False)
        self.assertEqual(detail.status_code, 303)
        batch = self.repository.get_crawl_batch(1)
        self.assertEqual(batch.success_count, 1)
        self.assertEqual(batch.pending_count, 1)
        client.post("/crawl-batches/1/run-next", data={"chunk_size": "20"})
        batch = self.repository.get_crawl_batch(1)
        self.assertEqual(batch.status, "completed")
        self.assertEqual(batch.success_count, 2)
        self.assertEqual(provider.calls, [("UC1", None), ("UC2", None)])

    def test_full_crawl_batch_failure_does_not_stop_and_retry_is_explicit(self) -> None:
        provider = BatchFailureFake()
        client = TestClient(create_app(repository=self.repository, video_provider=provider))
        client.post("/channels/full-crawl-batch", data={"limit": "2", "confirmation": "true"})
        client.post("/crawl-batches/1/run-next", data={"chunk_size": "5"})
        batch = self.repository.get_crawl_batch(1)
        self.assertEqual(batch.status, "completed_with_errors")
        self.assertEqual((batch.success_count, batch.failure_count), (1, 1))
        self.assertEqual(self.repository.list_crawl_batch_items(1)[0].status, "failed")
        client.post("/crawl-batches/1/retry-failed")
        self.assertEqual(self.repository.list_crawl_batch_items(1)[0].status, "pending")

    def test_crawl_batch_history_and_detail_are_read_only(self) -> None:
        self.assertEqual(self.client.get("/crawl-batches").status_code, 200)
        self.assertEqual(self.client.get("/crawl-batches/999").status_code, 404)

    def test_channels_default_pagination_and_valid_page_sizes(self) -> None:
        repo = TranscriptRepository(Path(self.temp.name) / "many.db")
        for number in range(81):
            channel_id = f"UC{number:03d}"
            repo.upsert_channel(Channel(channel_id, f"Channel {number:03d}"))
            repo.record_discovery(channel_id, "retirement", "seed")
        client = TestClient(create_app(repository=repo))
        first = client.get("/channels")
        self.assertIn("Page 1 of 2", first.text)
        self.assertEqual(first.text.count('href="/channels/UC'), 50)
        second = client.get("/channels?page=2")
        self.assertIn("Page 2 of 2", second.text)
        self.assertEqual(second.text.count('href="/channels/UC'), 31)
        self.assertNotIn(">Next<", second.text)
        self.assertIn("Previous", second.text)
        self.assertIn("Page 1 of 4", client.get("/channels?per_page=25").text)
        self.assertIn("Page 1 of 1", client.get("/channels?per_page=100").text)
        self.assertIn("Page 1 of 2", client.get("/channels?per_page=10").text)
        self.assertIn("Page 1 of 2", client.get("/channels?per_page=abc").text)

    def test_channels_pagination_preserves_filters_and_clamps_large_page(self) -> None:
        repo = TranscriptRepository(Path(self.temp.name) / "filtered-many.db")
        for number in range(81):
            channel_id = f"UF{number:03d}"
            repo.upsert_channel(Channel(channel_id, f"Filtered {number:03d}"))
            repo.record_discovery(channel_id, "Retirement Planning", "seed")
        client = TestClient(create_app(repository=repo))
        response = client.get("/channels?per_page=25&keyword= RETIREMENT   PLANNING &sort=subscribers")
        self.assertIn("Page 1 of 4", response.text)
        self.assertIn("/channels?per_page=25&amp;keyword=+RETIREMENT+++PLANNING+&amp;sort=subscribers&amp;page=2", response.text)
        clamped = client.get("/channels?page=999&per_page=25&keyword=retirement%20planning")
        self.assertIn("Page 4 of 4", clamped.text)
        self.assertNotIn(">Next<", clamped.text)

    def test_keyword_filter_normalizes_variants_and_count_matches_rows(self) -> None:
        for keyword in ("retirement", " RETIREMENT ", "retirement   "):
            response = self.client.get(f"/channels?keyword={keyword}")
            self.assertIn("One", response.text)
            self.assertIn("Two", response.text)
        self.assertEqual(self.repository.count_channels_page(keyword=" RETIREMENT "), 2)
        self.assertEqual(len(self.repository.list_channels_page(50, keyword="RETIREMENT")), 2)

    def test_channels_page_exposes_full_crawl_batch_action(self) -> None:
        response = self.client.get("/channels")
        self.assertIn('action="/channels/full-crawl-batch"', response.text)
        self.assertIn("Create Full Crawl Batch", response.text)
        self.assertNotIn('action="/channels/score-unscored"', response.text)
        self.assertIn('name="limit"', response.text)

    def test_score_unscored_requires_positive_limit_and_updates_scores(self) -> None:
        self.assertEqual(self.client.post("/channels/score-unscored", data={"limit": "0"}).status_code, 400)
        response = self.client.post("/channels/score-unscored", data={"limit": "1"}, follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        self.assertIn("/channels?message=", response.headers["location"])
        self.assertIsNotNone(self.repository.get_channel_score("UC1"))

    def test_detail_and_unknown_404(self) -> None:
        self.assertEqual(self.client.get("/channels/UC1").status_code, 200)
        self.assertEqual(self.client.get("/videos/v1").status_code, 200)
        self.assertEqual(self.client.get("/channels/missing").status_code, 404)
        self.assertEqual(self.client.get("/videos/missing").status_code, 404)

    def test_get_requests_do_not_create_work_plan(self) -> None:
        before = len(self.repository.list_work_plans(100))
        self.client.get("/work-plans")
        self.assertEqual(len(self.repository.list_work_plans(100)), before)

    def test_create_work_plan_post_and_validation(self) -> None:
        response = self.client.post(
            "/work-plans", data={"max_crawls": "0", "max_enrichments": "2", "max_transcripts": "1"}, follow_redirects=False
        )
        self.assertEqual(response.status_code, 303)
        self.assertIn("/work-plans/", response.headers["location"])
        invalid = self.client.post(
            "/work-plans", data={"max_crawls": "-1", "max_enrichments": "2", "max_transcripts": "1"}
        )
        self.assertEqual(invalid.status_code, 400)

    def test_execute_requires_explicit_max_items(self) -> None:
        plan = self.repository.list_work_plans(1)
        if not plan:
            self.repository.create_work_plan(
                __import__("src.crawl_yt.database.models", fromlist=["WorkPlan"]).WorkPlan(
                    None, __import__("datetime").datetime.now(__import__("datetime").timezone.utc), "planned",
                    __import__("src.crawl_yt.database.models", fromlist=["OperationalBudget"]).OperationalBudget(0, 0, 0),
                    {"crawl_channel": 0, "enrich_video": 0, "transcript_video": 0},
                ), []
            )
            plan = self.repository.list_work_plans(1)
        response = self.client.post(f"/work-plans/{plan[0].id}/execute", data={})
        self.assertEqual(response.status_code, 400)

    def test_full_crawl_requires_confirmation(self) -> None:
        response = self.client.post("/channels/UC1/crawl-full", data={})
        self.assertEqual(response.status_code, 400)

    def test_transcript_detail_renders_timestamped_segments(self) -> None:
        response = self.client.get("/videos/v1/transcripts/en/youtube_auto")
        self.assertEqual(response.status_code, 200)
        self.assertIn("00:00:12.400", response.text)
        self.assertIn("Police arrived.", response.text)
        self.assertIn("youtube_auto", response.text)

    def test_missing_transcript_detail_is_404(self) -> None:
        self.assertEqual(self.client.get("/videos/v1/transcripts/fr/youtube_auto").status_code, 404)
        self.assertEqual(self.client.get("/videos/missing/transcripts/en/youtube_auto").status_code, 404)

    def test_dashboard_uses_count_methods_not_due_materialization(self) -> None:
        self.repository.list_channels_due_for_crawl = lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("materialized due rows"))
        self.assertEqual(self.client.get("/").status_code, 200)

    def test_web_source_does_not_use_private_connect(self) -> None:
        import inspect
        from src.crawl_yt.web import app as web_app
        self.assertNotIn("._connect(", inspect.getsource(web_app))

    def test_discovery_get_has_no_provider_call_and_post_calls_fake(self) -> None:
        provider = DiscoveryFake()
        app_client = TestClient(create_app(repository=self.repository, discovery_provider=provider))
        self.assertEqual(app_client.get("/discovery").status_code, 200)
        self.assertEqual(provider.calls, [])
        self.assertEqual(app_client.post("/discovery", data={"keyword": "retirement", "limit": "1", "related_terms": "planning"}).status_code, 200)
        self.assertEqual(provider.calls, [("retirement", 100)])

    def test_discovery_result_is_structured_and_labels_statuses(self) -> None:
        provider = DiscoveryFake()
        response = TestClient(create_app(repository=self.repository, discovery_provider=provider)).post("/discovery", data={"keyword": "retirement", "limit": "20", "related_terms": "planning"})
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("DiscoveryReport(", response.text)
        for label in ("Search results", "Unique channels", "Duplicate results in search", "New channels", "Existing channels", "New discovery relationships", "Existing discovery relations"):
            self.assertIn(label, response.text)
        self.assertIn("Two", response.text)
        self.assertIn("/channels/UC3", response.text)
        self.assertIn("https://youtube.com/channel/UC3", response.text)
        self.assertIn("target=\"_blank\"", response.text)
        self.assertIn("New", response.text)
        self.assertIn("Existing", response.text)
        self.assertIn("Channels scored", response.text)
        self.assertIn("Scoring failures", response.text)
        self.assertIn("/channels?keyword=retirement", response.text)

    def test_discovery_result_renders_query_metrics_and_candidate_provenance(self) -> None:
        profile = self.repository.create_topic_profile(
            "Retirement", "", ["retirement planning"], ["retirement income", "broken query"],
        )
        response = TestClient(create_app(
            repository=self.repository, discovery_provider=MultiQueryDiscoveryFake(),
        )).post(
            "/discovery",
            data={"keyword": "retirement", "limit": "5", "topic_profile_id": str(profile.id)},
        )

        self.assertEqual(response.status_code, 200)
        for label in (
            "Planned queries", "Executed queries", "Raw results", "Unique candidates",
            "Cross-query duplicates", "Query breakdown", "Found by",
        ):
            self.assertIn(label, response.text)
        for metric in (
            '<span>Planned queries</span><strong>3</strong>',
            '<span>Executed queries</span><strong>3</strong>',
            '<span>Raw results</span><strong>4</strong>',
            '<span>Unique candidates</span><strong>3</strong>',
            '<span>Cross-query duplicates</span><strong>1</strong>',
            '<span>Accepted</span><strong>2</strong>',
            '<span>Rejected</span><strong>1</strong>',
        ):
            self.assertIn(metric, response.text)
        self.assertIn("retirement income", response.text)
        self.assertIn("<td>2</td><td>2</td><td>1</td><td>1</td>", response.text)
        self.assertIn("search unavailable", response.text)
        self.assertIn("retirement; retirement income", response.text)
        self.assertIn("<td>retirement</td><td>Coverage below Balanced minimum.</td>", response.text)
        self.assertNotIn("DiscoveryReport(", response.text)
        self.assertNotIn("DiscoveryQueryMetric(", response.text)

    def test_empty_discovery_result_is_friendly(self) -> None:
        response = TestClient(create_app(repository=self.repository, discovery_provider=EmptyDiscoveryFake())).post("/discovery", data={"keyword": "unknown", "limit": "20"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("No channels found for this keyword.", response.text)
        self.assertNotIn("DiscoveryReport(", response.text)

    def test_discovery_dry_run_notice_and_discovered_status(self) -> None:
        response = TestClient(create_app(repository=self.repository, discovery_provider=DiscoveryFake())).post("/discovery", data={"keyword": "retirement", "limit": "20", "related_terms": "planning", "dry_run": "true"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("Dry run — nothing was written to the database.", response.text)
        self.assertIn("Discovered", response.text)

    def test_discovery_relevance_ui_shows_mode_related_terms_and_rejections(self) -> None:
        response = TestClient(create_app(repository=self.repository, discovery_provider=RelevanceDiscoveryFake())).post(
            "/discovery", data={"keyword": "retirement", "limit": "2", "mode": "strict", "related_terms": "planning, PLANNING"}
        )
        self.assertEqual(response.status_code, 200)
        for label in ("Candidates found", "Accepted", "Rejected", "Mode", "Minimum coverage", "Minimum distinct concepts", "Distinct concepts", "Strong identity floor", "Accepted", "Rejected candidates", "Topic coverage", "Matched sample", "Identity", "Reason", "Evidence"):
            self.assertIn(label, response.text)
        self.assertIn("Extra concepts: planning", response.text)
        self.assertIn("Accepted", response.text)
        self.assertIn("1 / 20", response.text)
        self.assertIn("Rejected<br>", response.text)

    def test_discovery_form_explains_target_accepted_semantics(self) -> None:
        response = self.client.get("/discovery")
        self.assertIn("Target accepted channels", response.text)
        self.assertIn("Discovery may inspect more candidates to find this many relevant channels.", response.text)

    def test_discovery_mode_propagates_through_web_boundary(self) -> None:
        for mode, matches, accepted in (("strict", 6, 0), ("balanced", 6, 1), ("broad", 6, 1), ("strict", 9, 1), ("balanced", 9, 1), ("broad", 9, 1), ("strict", 13, 1), ("balanced", 13, 1), ("broad", 13, 1)):
            provider = ModeDiscoveryFake(matches)
            response = TestClient(create_app(repository=self.repository, discovery_provider=provider)).post(
                "/discovery", data={"keyword": "retirement", "limit": "1", "mode": mode, "related_terms": "living alone"}
            )
            self.assertEqual(response.status_code, 200)
            self.assertIn(f"<strong>{mode.title()}</strong>", response.text)
            self.assertIn(f"<span>Accepted</span><strong>{accepted}</strong>", response.text)
            minimum = {"strict": 40, "balanced": 20, "broad": 10}[mode]
            identity_floor = {"strict": 30, "balanced": 15, "broad": 5}[mode]
            self.assertIn(f"<span>Minimum coverage</span><strong>{minimum}%</strong>", response.text)
            self.assertIn("<span>Minimum distinct concepts</span><strong>2</strong>", response.text)
            self.assertIn(f"<span>Strong identity floor</span><strong>{identity_floor}%</strong>", response.text)
            self.assertEqual(provider.calls, [("retirement", 100)])

    def test_discovery_result_shows_target_and_rejection_buckets(self) -> None:
        provider = ModeDiscoveryFake(0)
        response = TestClient(create_app(repository=self.repository, discovery_provider=provider)).post(
            "/discovery", data={"keyword": "retirement", "limit": "20", "mode": "balanced"}
        )
        for label in ("Target accepted", "Maximum candidates", "Raw results", "Unique candidates", "No usable video sample", "0–15% coverage", "Verification failed"):
            self.assertIn(label, response.text)
        self.assertIn("Candidates were mostly off-topic. Consider adding extra concepts or using Broad.", response.text)

    def test_discovery_empty_sample_diagnostic_message(self) -> None:
        response = TestClient(create_app(repository=self.repository, discovery_provider=EmptySampleDiscoveryFake())).post(
            "/discovery", data={"keyword": "retirement", "limit": "1", "mode": "balanced"}
        )
        self.assertIn("No usable video sample: 1", response.text)
        self.assertIn("Channel topic verification could not obtain enough recent video titles.", response.text)

    def test_topic_profile_crud_ui_and_discovery_selector(self) -> None:
        created = self.client.post(
            "/topic-profiles",
            data={
                "name": "Solo Aging",
                "description": "Independent later life",
                "concept_phrases": "Living Alone\nliving   alone\nSenior Life",
                "search_concepts": "solo retirement\naging independently",
            },
            follow_redirects=False,
        )
        self.assertEqual(created.status_code, 303)
        detail_url = created.headers["location"]
        detail = self.client.get(detail_url)
        self.assertIn("Solo Aging", detail.text)
        self.assertIn("living alone", detail.text)
        self.assertIn("senior life", detail.text)
        self.assertIn("Verification terms", detail.text)
        self.assertIn("Candidate-finding terms", detail.text)
        self.assertIn("solo retirement", detail.text)
        self.assertIn("aging independently", detail.text)
        self.assertIn("Solo Aging", self.client.get("/topic-profiles").text)
        discovery = self.client.get("/discovery")
        self.assertIn("Topic profile", discovery.text)
        self.assertIn("Extra concepts", discovery.text)
        self.assertIn('class="target-field"', discovery.text)

        profile_id = int(detail_url.rsplit("/", 1)[-1])
        updated = self.client.post(
            f"/topic-profiles/{profile_id}",
            data={
                "name": "Solo Aging Updated",
                "description": "",
                "concept_phrases": "independent aging",
                "search_concepts": "later life planning",
            },
            follow_redirects=False,
        )
        self.assertEqual(updated.status_code, 303)
        updated_detail = self.client.get(detail_url).text
        self.assertIn("Solo Aging Updated", updated_detail)
        self.assertIn("independent aging", updated_detail)
        self.assertIn("later life planning", updated_detail)
        deleted = self.client.post(f"/topic-profiles/{profile_id}/delete", follow_redirects=False)
        self.assertEqual(deleted.status_code, 303)
        self.assertNotIn("Solo Aging Updated", self.client.get("/topic-profiles").text)

    def test_legacy_topic_profile_renders_without_candidate_finding_terms(self) -> None:
        profile = self.repository.create_topic_profile("Legacy", "", ["living alone"])

        detail = self.client.get(f"/topic-profiles/{profile.id}")
        edit = self.client.get(f"/topic-profiles/{profile.id}/edit")
        listing = self.client.get("/topic-profiles")

        self.assertEqual(detail.status_code, 200)
        self.assertIn("Verification terms", detail.text)
        self.assertIn("living alone", detail.text)
        self.assertIn("No candidate-finding terms.", detail.text)
        self.assertIn('name="search_concepts"', edit.text)
        self.assertIn("Candidate-finding terms", edit.text)
        self.assertIn("Candidate-finding terms", listing.text)

    def test_legacy_profile_update_preserves_omitted_candidate_finding_terms(self) -> None:
        profile = self.repository.create_topic_profile(
            "Solo Aging", "", ["living alone"], ["solo retirement"],
        )

        updated = self.client.post(
            f"/topic-profiles/{profile.id}",
            data={"name": "Solo Aging", "description": "", "concept_phrases": "aging alone"},
            follow_redirects=False,
        )

        self.assertEqual(updated.status_code, 303)
        detail = self.client.get(f"/topic-profiles/{profile.id}").text
        self.assertIn("aging alone", detail)
        self.assertIn("solo retirement", detail)

    def test_discovery_profile_renders_snapshot_top_concepts_and_title_evidence(self) -> None:
        profile = self.repository.create_topic_profile("Solo Aging", "", ["living alone", "senior life"])
        provider = ProfileDiscoveryFake()
        response = TestClient(create_app(repository=self.repository, discovery_provider=provider)).post(
            "/discovery",
            data={"keyword": "solo aging", "limit": "1", "mode": "balanced", "topic_profile_id": str(profile.id), "extra_concepts": "independent aging"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("Topic profile", response.text)
        self.assertIn("Solo Aging", response.text)
        self.assertIn("Concepts used", response.text)
        self.assertIn("Top matched concepts", response.text)
        self.assertIn("living alone", response.text)
        self.assertIn("Why I enjoy living alone after 65", response.text)
        self.assertIn("matched: living alone", response.text)
        self.assertIn("Kitchen tools", response.text)
        self.assertEqual(provider.search_calls, [("solo aging", 100)])
        self.assertEqual(provider.verify_calls, [("UCPROFILE", 20)])

    def test_discovery_history_shows_keyword_counts_and_links(self) -> None:
        response = self.client.get("/discovery")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Discovery History", response.text)
        self.assertIn("retirement", response.text)
        self.assertIn("/discovery/keywords/retirement", response.text)
        self.assertIn(">2<", response.text)

    def test_keyword_detail_is_paginated_and_associated_only(self) -> None:
        response = self.client.get("/discovery/keywords/retirement?page=1&per_page=1")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Discovery: retirement", response.text)
        self.assertIn("Total channels", response.text)
        self.assertIn("One", response.text)
        self.assertIn("Page 1", response.text)
        self.assertIn("2 total", response.text)

    def test_keyword_detail_shows_other_keywords_and_unicode(self) -> None:
        response = self.client.get("/discovery/keywords/social%20security")
        self.assertEqual(response.status_code, 200)
        self.assertIn("One", response.text)
        self.assertIn("retirement", response.text)

    def test_channels_keyword_filter_and_provenance(self) -> None:
        response = self.client.get("/channels?keyword=retirement")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Discovery keywords", response.text)
        self.assertIn("One", response.text)
        self.assertIn("Two", response.text)
        self.assertIn("retirement", response.text)

    def test_channel_detail_links_provenance(self) -> None:
        response = self.client.get("/channels/UC1")
        self.assertIn("/discovery/keywords/retirement", response.text)
        self.assertIn("/discovery/keywords/social%20security", response.text)

    def test_unscored_channel_is_human_readable(self) -> None:
        response = self.client.get("/channels/UC1")
        self.assertIn("Not scored yet", response.text)
        self.assertIn("Never crawled", response.text)
        self.assertNotIn("<pre>None</pre>", response.text)

    def test_scored_channel_renders_structured_score_and_reasons(self) -> None:
        self.repository.upsert_channel_score(ChannelScore("UC1", 82.4, 91.0, 78.5, 74.2, 85.0, "high", {"notes": ["active uploader"], "videos_per_week_30d": 5.8, "videos_per_week_90d": 5.4, "observation_coverage_30d": True, "cadence_fit": "very good", "consistency": "high", "score_maturity": "mature"}, datetime.now(timezone.utc), "v2", 78.5, 5.8, 5.4))
        response = self.client.get("/channels/UC1")
        for label in ("Overall Score", "82.4", "HIGH", "Relevance", "Upload cadence", "Traction", "Confidence", "Scoring version", "v2", "5.8 videos/week", "5.4 videos/week", "VERY GOOD", "HIGH", "active uploader"):
            self.assertIn(label, response.text)
        self.assertNotIn("ChannelScore(", response.text)

    def test_channels_table_supports_cadence_filter_and_sort(self) -> None:
        now = datetime.now(timezone.utc)
        self.repository.upsert_channel_score(ChannelScore("UC1", 70, 70, 80, 60, 60, "high", {"observation_coverage_30d": True, "cadence_fit": "very good"}, now, "v2", 80, 5.8, 5.0))
        self.repository.upsert_channel_score(ChannelScore("UC2", 80, 80, 60, 70, 70, "high", {"observation_coverage_30d": True, "cadence_fit": "below target"}, now, "v2", 60, 2.0, 2.0))
        filtered = self.client.get("/channels?min_videos_per_week=5").text
        self.assertIn("One", filtered)
        self.assertNotIn('href="/channels/UC2"', filtered)
        sorted_page = self.client.get("/channels?sort=cadence").text
        self.assertLess(sorted_page.index("One"), sorted_page.index("Two"))
        self.assertIn("Videos/week", sorted_page)
        self.assertIn("5.8 / week", sorted_page)
        self.assertIn("Cadence fit", sorted_page)

    def test_channels_table_hides_insufficient_cadence_coverage(self) -> None:
        now = datetime.now(timezone.utc)
        self.repository.upsert_channel_score(ChannelScore("UC1", 70, 70, 80, 60, 60, "high", {"videos_per_week_30d": 5.8, "observation_coverage_30d": False}, now, "v2", 80, 5.8, 5.0))
        response = self.client.get("/channels?min_videos_per_week=5")
        self.assertNotIn('href="/channels/UC1"', response.text)
        self.assertIn("Not enough data", self.client.get("/channels").text)

    def test_score_post_redirects_and_score_appears(self) -> None:
        response = self.client.post("/channels/UC1/score", follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        detail = self.client.get(response.headers["location"])
        self.assertIn("Overall Score", detail.text)

    def test_crawl_state_and_metadata_empty_states_are_human_readable(self) -> None:
        self.repository.ensure_channel_crawl_state("UC1")
        self.repository.mark_crawl_success("UC1", "v1", None, datetime(2026, 8, 21, 17, 46, tzinfo=timezone.utc))
        response = self.client.get("/channels/UC1")
        self.assertIn("Last success", response.text)
        self.assertIn("2026-08-21 17:46 UTC", response.text)
        self.assertNotIn("ChannelCrawlState(", response.text)

    def test_channel_detail_shows_policy_interval_for_unscored(self) -> None:
        response = self.client.get("/channels/UC1")
        self.assertIn("Recommended interval", response.text)
        self.assertIn("1 day", response.text)
        self.assertNotIn("1 day, 0:00:00", response.text)

    def test_channel_detail_shows_policy_interval_by_tier(self) -> None:
        now = datetime.now(timezone.utc)
        for channel_id, tier, interval in (("UC1", "high", "3 days"), ("UC2", "medium", "7 days")):
            self.repository.upsert_channel_score(ChannelScore(channel_id, 80, 80, 80, 80, 80, tier, {}, now, "v2"))
            response = self.client.get(f"/channels/{channel_id}")
            self.assertIn(interval, response.text)
            self.assertNotIn("days, 0:00:00", response.text)
        self.repository.upsert_channel(Channel("UC3", "Three"))
        self.repository.upsert_channel_score(ChannelScore("UC3", 20, 20, 20, 20, 20, "low", {}, now, "v2"))
        self.assertIn("14 days", self.client.get("/channels/UC3").text)

    def test_channels_table_shows_recommended_intervals(self) -> None:
        now = datetime.now(timezone.utc)
        self.repository.upsert_channel_score(ChannelScore("UC1", 80, 80, 80, 80, 80, "high", {}, now, "v2"))
        self.assertIn("3 days", self.client.get("/channels").text)
        self.assertIn("1 day", self.client.get("/channels").text)

    def test_channel_export_defaults_to_score_60_and_formats_workbook(self) -> None:
        now = datetime.now(timezone.utc)
        self.repository.upsert_channel(Channel("UC59", "Below", channel_url="https://example.com/below"))
        self.repository.upsert_channel(Channel("UC60", "Included", channel_url="https://example.com/included"))
        self.repository.upsert_channel(Channel("UC80", "High"))
        self.repository.update_channel_metadata(ChannelMetadata("UC60", subscriber_count=1234, view_count=5678, video_count=99))
        self.repository.record_discovery("UC60", "retirement", "seed", now)
        for channel_id, score in (("UC59", 59.9), ("UC60", 60), ("UC80", 80)):
            self.repository.upsert_channel_score(ChannelScore(channel_id, score, 70, 80, 75, 60, "high", {"videos_per_week_30d": 5.5, "videos_per_week_90d": 4.2, "cadence_fit": "very good", "median_enriched_views": 40000, "view_subscriber_ratio": 0.8, "score_maturity": "mature"}, now, "v2", 85, 5.5, 4.2))
        response = self.client.get("/channels/export.xlsx")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["content-type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("attachment; filename=\"crawl-yt-channels-score-60.xlsx\"", response.headers["content-disposition"])
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["Channels"]
        self.assertEqual([cell.value for cell in sheet[1]], ["Channel", "Channel URL", "Channel ID", "Overall Score", "Tier", "Videos/Week 30d", "Videos/Week 90d", "Cadence Fit", "Subscribers", "Channel Views", "Reported Video Count", "Median Video Views", "Views / Subscribers", "Discovery Keywords", "Score Maturity", "Scoring Version", "Last Crawl", "Next Crawl"])
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertIsNotNone(sheet.auto_filter.ref)
        values = list(sheet.iter_rows(min_row=2, values_only=False))
        names = [row[0].value for row in values]
        self.assertEqual(names, ["High", "Included"])
        self.assertEqual(values[1][1].value, "https://example.com/included")
        self.assertEqual(values[1][1].hyperlink.target, "https://example.com/included")
        self.assertEqual(values[1][8].value, 1234)
        self.assertEqual(values[1][9].value, 5678)
        self.assertEqual(values[1][10].value, 99)
        self.assertEqual(values[1][13].value, "retirement")
        self.assertEqual(values[0][13].value, None)

    def test_channel_export_respects_filters_and_ignores_pagination(self) -> None:
        now = datetime.now(timezone.utc)
        repo = TranscriptRepository(Path(self.temp.name) / "export-many.db")
        for number in range(81):
            channel_id = f"EX{number:03d}"
            repo.upsert_channel(Channel(channel_id, f"Export {number:03d}"))
            repo.record_discovery(channel_id, "retirement", "seed", now)
            repo.upsert_channel_score(ChannelScore(channel_id, 80, 70, 80, 75, 60, "high", {"observation_coverage_30d": True, "cadence_fit": "very good"}, now, "v2", 85, 5.5, 4.2))
        client = TestClient(create_app(repository=repo))
        response = client.get("/channels/export.xlsx?min_score=60&keyword=RETIREMENT&min_videos_per_week=3&tier=high&page=2&per_page=50")
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertEqual(workbook["Channels"].max_row, 82)

    def test_channel_export_empty_and_safety_limit_are_friendly(self) -> None:
        self.assertEqual(self.client.get("/channels/export.xlsx?min_score=100").status_code, 400)
        self.repository.list_channels_for_export = lambda **kwargs: (_ for _ in ()).throw(ValueError("Export matches more than 50,000 channels. Narrow your filters."))
        response = self.client.get("/channels/export.xlsx")
        self.assertEqual(response.status_code, 400)
        self.assertIn("50,000", response.text)

    def test_optional_numeric_channel_filters_accept_blank_values(self) -> None:
        for query in ("min_score=", "min_videos_per_week=", "min_score=&min_videos_per_week="):
            response = self.client.get(f"/channels?{query}")
            self.assertEqual(response.status_code, 200)
        response = self.client.get("/channels?min_score=abc")
        self.assertEqual(response.status_code, 400)
        self.assertIn("minimum score must be a number", response.text)
        self.assertNotIn('"detail"', response.text)
        response = self.client.get("/channels?min_videos_per_week=abc")
        self.assertEqual(response.status_code, 400)
        self.assertIn("minimum videos/week must be a number", response.text)

    def test_optional_numeric_channel_filters_validate_ranges_and_decimals(self) -> None:
        for query in ("min_score=60", "min_score=60.5", "min_videos_per_week=3", "min_videos_per_week=3.5"):
            self.assertEqual(self.client.get(f"/channels?{query}").status_code, 200)
        for query in ("min_score=101", "min_score=-1", "min_videos_per_week=-1"):
            self.assertEqual(self.client.get(f"/channels?{query}").status_code, 400)

    def test_channel_pagination_urls_omit_blank_numeric_filters(self) -> None:
        for number in range(51):
            self.repository.upsert_channel(Channel(f"PAGE{number:02d}", f"Page {number:02d}"))
        response = self.client.get("/channels?page=2&min_score=&min_videos_per_week=&keyword=retirement")
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("min_score=", response.text)
        self.assertNotIn("min_videos_per_week=", response.text)

    def test_export_blank_score_defaults_to_60_and_blank_cadence_is_ignored(self) -> None:
        captured: list[dict[str, object]] = []
        self.repository.list_channels_for_export = lambda **kwargs: (captured.append(kwargs) or [])
        self.assertEqual(self.client.get("/channels/export.xlsx?min_score=&min_videos_per_week=").status_code, 400)
        self.assertEqual(captured[-1]["min_score"], 60)
        self.assertIsNone(captured[-1]["min_videos_per_week"])
        self.assertEqual(self.client.get("/channels/export.xlsx?min_score=70").status_code, 400)
        self.assertEqual(captured[-1]["min_score"], 70)

    def test_channel_zero_subscribers_is_not_missing(self) -> None:
        self.repository.upsert_channel(Channel("UC0", "Zero", subscriber_count=0))
        response = self.client.get("/channels/UC0")
        self.assertIn("<dt>Subscribers</dt><dd>0</dd>", response.text)

    def test_channel_metadata_fields_and_large_numbers_are_human_readable(self) -> None:
        self.repository.update_channel_metadata(ChannelMetadata("UC1", subscriber_count=1250000, view_count=9876543, video_count=42))
        response = self.client.get("/channels/UC1")
        self.assertIn("1,250,000", response.text)
        self.assertIn("9,876,543", response.text)
        self.assertIn("Reported videos", response.text)

    def test_channels_source_links_prefer_persisted_url_and_fallback_without_network(self) -> None:
        self.repository.upsert_channel(Channel("UCURL", "With URL", channel_url="https://www.youtube.com/@with-url"))
        response = self.client.get("/channels")
        self.assertIn('href="https://www.youtube.com/@with-url" target="_blank" rel="noopener noreferrer">YouTube', response.text)
        self.assertIn('href="https://www.youtube.com/channel/UC1" target="_blank" rel="noopener noreferrer">YouTube', response.text)
        self.assertIn('href="/channels/UC1">One</a>', response.text)

    def test_channel_detail_source_and_metadata_timestamp_are_human_readable(self) -> None:
        checked = datetime(2026, 8, 22, 10, 2, tzinfo=timezone.utc)
        self.repository.update_channel_metadata(ChannelMetadata("UC1", checked_at=checked))
        response = self.client.get("/channels/UC1")
        self.assertIn("Channel ID:", response.text)
        self.assertIn("Open on YouTube", response.text)
        self.assertIn('href="https://www.youtube.com/channel/UC1" target="_blank" rel="noopener noreferrer"', response.text)
        self.assertIn("2026-08-22 10:02 UTC", response.text)
        unknown = self.client.get("/channels/UC2")
        self.assertIn("Metadata checked", unknown.text)
        self.assertIn("—", unknown.text)

    def test_channel_without_videos_has_empty_state(self) -> None:
        self.repository.upsert_channel(Channel("UC0", "Empty"))
        response = self.client.get("/channels/UC0")
        self.assertIn("No videos stored yet.", response.text)

    def test_video_score_is_structured(self) -> None:
        score = VideoScore("v1", 77.5, 80.0, 75.0, 70.0, 65.0, 85.0, 78.0, "high", '{"metadata_priority": 70, "transcript_priority": 80}', datetime.now(timezone.utc), "v1")
        self.repository.upsert_video_score(score)
        response = self.client.get("/videos/v1")
        self.assertIn("Overall priority", response.text)
        self.assertIn("77.5", response.text)
        self.assertIn("Metadata priority", response.text)
        self.assertNotIn("VideoScore(", response.text)

    def test_empty_discovery_history_and_keyword_states(self) -> None:
        empty_repo = TranscriptRepository(Path(self.temp.name) / "empty.db")
        empty_repo.upsert_channel(Channel("UC9", "No provenance"))
        empty_client = TestClient(create_app(repository=empty_repo))
        self.assertIn("No discovery searches have been saved yet.", empty_client.get("/discovery").text)
        self.assertIn("No channels are currently associated with this keyword.", empty_client.get("/discovery/keywords/unknown").text)

    def test_channel_crawl_post_calls_fake_provider(self) -> None:
        provider = VideoFake(Video("v3", "UC1", "New video", datetime.now(timezone.utc)))
        app_client = TestClient(create_app(repository=self.repository, video_provider=provider))
        response = app_client.post("/channels/UC1/crawl", data={}, follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        self.assertEqual(provider.calls, [("UC1", None)])

    def test_video_actions_call_fake_services(self) -> None:
        metadata = MetadataFake()
        transcript = TranscriptFake()
        app_client = TestClient(create_app(repository=self.repository, metadata_provider=metadata, transcript_provider=transcript))
        self.assertEqual(app_client.post("/videos/v1/score", follow_redirects=False).status_code, 303)
        self.assertEqual(app_client.post("/videos/v2/enrich", follow_redirects=False).status_code, 303)
        self.assertEqual(app_client.post("/videos/v2/transcript", data={"language": "en"}, follow_redirects=False).status_code, 303)
        self.assertEqual(metadata.calls[0][0], "v2")
        self.assertEqual(transcript.calls[0][2], ("en",))

    def test_unknown_action_targets_are_404(self) -> None:
        self.assertEqual(self.client.post("/channels/missing/crawl", data={}).status_code, 404)
        self.assertEqual(self.client.post("/videos/missing/score").status_code, 404)
        self.assertEqual(self.client.post("/videos/missing/enrich").status_code, 404)
        self.assertEqual(self.client.post("/videos/missing/transcript", data={}).status_code, 404)
        self.assertEqual(self.client.get("/work-plans/999999").status_code, 404)
        self.assertEqual(self.client.post("/work-plans/999999/execute", data={"max_items": "1"}).status_code, 404)

    def test_provider_failure_is_friendly_html(self) -> None:
        class FailingVideoProvider:
            def iterate_videos(self, channel_id, limit=None):
                raise RuntimeError("provider unavailable")
        response = TestClient(create_app(repository=self.repository, video_provider=FailingVideoProvider()), raise_server_exceptions=False).post("/channels/UC1/crawl", data={})
        self.assertEqual(response.status_code, 502)
        self.assertIn("Operation failed", response.text)

    def test_transcript_action_is_caption_only(self) -> None:
        transcript = TranscriptFake()
        self.assertEqual(TestClient(create_app(repository=self.repository, transcript_provider=transcript)).post("/videos/v2/transcript", data={}, follow_redirects=False).status_code, 303)
        self.assertEqual(transcript.calls[0][2], ("en", "en-US", "en-GB"))

    def test_work_plan_execution_uses_mocked_provider(self) -> None:
        metadata = MetadataFake()
        app_client = TestClient(create_app(repository=self.repository, metadata_provider=metadata))
        created = app_client.post("/work-plans", data={"max_crawls": "0", "max_enrichments": "1", "max_transcripts": "0"}, follow_redirects=False)
        self.assertEqual(created.status_code, 303)
        plan_id = created.headers["location"].rsplit("/", 1)[-1]
        executed = app_client.post(f"/work-plans/{plan_id}/execute", data={"max_items": "1"}, follow_redirects=False)
        self.assertEqual(executed.status_code, 303)
        self.assertTrue(metadata.calls)


if __name__ == "__main__":
    unittest.main()
